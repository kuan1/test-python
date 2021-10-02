# import numpy

# print(numpy.version.version)

# lst = [1,2,3,4,5]
# print(numpy.array(lst) == lst)

from os import path
from openpyxl import load_workbook

temp_xlsx = path.join(path.dirname(__file__), 'temp.xlsx')

wb = load_workbook(filename=temp_xlsx)

for sheet in wb.sheetnames:
  print(sheet)
  ws = wb[sheet]

  rows = ws.rows
  columns = ws.columns

  for idx,row in enumerate(rows):
    line = [col.value for col in row]
    print(line)
